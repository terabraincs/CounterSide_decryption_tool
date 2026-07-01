#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import re
import struct
import traceback
from contextlib import suppress
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from command import (
    STAGE2_FORMAT_DIRS,
    WRITE_MODES,
    best_decode_asset_name,
    clean_excel_value,
    clear_directory_contents,
    detect_language,
    ensure_base_dirs,
    ensure_stage2_dirs,
    iter_flatten_path_value,
    normalize_write_mode,
    output_path_for_write_mode,
    project_root_from_script,
    relative_posix,
    safe_source_folder_name,
    sanitize_filename,
    to_lua_literal,
    write_json,
    write_stage2_readme,
)

try:
    import UnityPy  # type: ignore
except Exception:
    UnityPy = None

try:
    from openpyxl import Workbook  # type: ignore
except Exception:
    Workbook = None

SCRIPT_VERSION = "2026-06-29-flat-source-folders"
RAW_CANDIDATE_EXTS = {".luac", ".lua", ".bytes", ".txt", ".text", ".textasset"}
CRYPTO2_MASKS = [
    14003937370121879411,
    295159725236528685,
    14656252856989855980,
    3126201044280739051,
    6176412274767465921,
    8501111619623644353,
    1001882303165547266,
    889784367385610816,
    8403001398375820177,
    15646421979254498160,
    15540104736269140030,
    4473111575030559303,
    16641115610173278858,
    7005653296469604124,
    7641466651897675454,
    18242667629599333687,
]
ODD_MASK = 0x5555555555555555
EVEN_MASK = 0xAAAAAAAAAAAAAAAA

OPNAMES = [
    "MOVE", "LOADI", "LOADF", "LOADK", "LOADKX", "LOADFALSE", "LFALSESKIP",
    "LOADTRUE", "LOADNIL", "GETUPVAL", "SETUPVAL", "GETTABUP", "GETTABLE",
    "GETI", "GETFIELD", "SETTABUP", "SETTABLE", "SETI", "SETFIELD",
    "NEWTABLE", "SELF", "ADDI", "ADDK", "SUBK", "MULK", "MODK", "POWK",
    "DIVK", "IDIVK", "BANDK", "BORK", "BXORK", "SHRI", "SHLI", "ADD",
    "SUB", "MUL", "MOD", "POW", "DIV", "IDIV", "BAND", "BOR", "BXOR",
    "SHL", "SHR", "MMBIN", "MMBINI", "MMBINK", "UNM", "BNOT", "NOT",
    "LEN", "CONCAT", "CLOSE", "TBC", "JMP", "EQ", "LT", "LE", "EQK",
    "EQI", "LTI", "LEI", "GTI", "GEI", "TEST", "TESTSET", "CALL",
    "TAILCALL", "RETURN", "RETURN0", "RETURN1", "FORLOOP", "FORPREP",
    "TFORPREP", "TFORCALL", "TFORLOOP", "SETLIST", "CLOSURE", "VARARG",
    "VARARGPREP", "EXTRAARG",
]

STRING_ASSET_TOKENS = {
    "STRING", "TEXT", "DESC", "DESCRIPTION", "TOOLTIP", "NAME", "VOICE",
    "SUBTITLE", "MAIL", "TITLE", "FILTER", "STORY", "CUTSCENE",
}


@dataclass
class LuaTable:
    array: dict[int, Any]
    fields: dict[Any, Any]

    def __init__(self) -> None:
        self.array = {}
        self.fields = {}

    def set_array(self, index: int, value: Any) -> None:
        self.array[index] = value

    def set_field(self, key: Any, value: Any) -> None:
        if isinstance(key, int) and key >= 1:
            self.array[key] = value
        else:
            self.fields[key] = value

    def to_python(self) -> Any:
        if not self.fields and self.array:
            keys = sorted(self.array)
            if keys == list(range(1, len(keys) + 1)):
                return [to_python(self.array[i]) for i in keys]
        result: dict[str, Any] = {}
        for key in sorted(self.array):
            result[str(key)] = to_python(self.array[key])
        for key, value in self.fields.items():
            result[str(key)] = to_python(value)
        return result


def to_python(value: Any) -> Any:
    if isinstance(value, LuaTable):
        return value.to_python()
    return value


class Reader:
    def __init__(self, data: bytes) -> None:
        self.data = data
        self.o = 0

    def read(self, n: int) -> bytes:
        if self.o + n > len(self.data):
            raise EOFError(f"파일 끝을 넘어서 읽으려 했습니다: offset={self.o}, size={n}, len={len(self.data)}")
        out = self.data[self.o:self.o + n]
        self.o += n
        return out

    def byte(self) -> int:
        return self.read(1)[0]

    def u32(self) -> int:
        return struct.unpack("<I", self.read(4))[0]

    def i64(self) -> int:
        return struct.unpack("<q", self.read(8))[0]

    def f64(self) -> float:
        return struct.unpack("<d", self.read(8))[0]

    def load_unsigned(self, limit: int = (1 << 63) - 1) -> int:
        value = 0
        while True:
            byte = self.byte()
            value = (value << 7) | (byte & 0x7F)
            if byte & 0x80:
                if value > limit:
                    raise ValueError(f"정수값이 너무 큽니다: {value}")
                return value

    def load_int(self) -> int:
        return self.load_unsigned((1 << 31) - 1)

    def load_string(self) -> str | None:
        size = self.load_unsigned()
        if size == 0:
            return None
        return self.read(size - 1).decode("utf-8", errors="replace")


def parse_header(r: Reader) -> dict[str, Any]:
    sig = r.read(4)
    if sig != b"\x1bLua":
        raise ValueError("Lua bytecode signature not found")
    version = r.byte()
    fmt = r.byte()
    luac_data = r.read(6)
    sizes = {
        "instruction": r.byte(),
        "lua_integer": r.byte(),
        "lua_number": r.byte(),
    }
    luac_int = r.i64()
    luac_num = r.f64()
    return {
        "signature": sig.hex(),
        "version_byte": version,
        "version": "Lua 5.4" if version == 0x54 else f"unknown({version:#x})",
        "format": fmt,
        "luac_data": luac_data.hex(),
        "sizes": sizes,
        "luac_int": luac_int,
        "luac_num": luac_num,
    }


def parse_constant(r: Reader) -> Any:
    tag = r.byte()
    if tag == 0:
        return None
    if tag == 1:
        return False
    if tag == 17:
        return True
    if tag == 3:
        return r.i64()
    if tag == 19:
        return r.f64()
    if tag in (4, 20):
        return r.load_string()
    raise ValueError(f"unknown Lua constant tag: {tag}")


def parse_proto(r: Reader, parent_source: str | None = None) -> dict[str, Any]:
    source = r.load_string() or parent_source
    linedefined = r.load_int()
    lastlinedefined = r.load_int()
    numparams = r.byte()
    is_vararg = r.byte()
    maxstacksize = r.byte()

    code_size = r.load_int()
    code = [r.u32() for _ in range(code_size)]

    const_size = r.load_int()
    constants = [parse_constant(r) for _ in range(const_size)]

    upvalue_size = r.load_int()
    upvalues = [(r.byte(), r.byte(), r.byte()) for _ in range(upvalue_size)]

    proto_size = r.load_int()
    protos = [parse_proto(r, source) for _ in range(proto_size)]

    lineinfo_size = r.load_int()
    r.read(lineinfo_size)

    abslineinfo_size = r.load_int()
    abslineinfo = [(r.load_int(), r.load_int()) for _ in range(abslineinfo_size)]

    locvar_size = r.load_int()
    locvars = [(r.load_string(), r.load_int(), r.load_int()) for _ in range(locvar_size)]

    upvalue_name_size = r.load_int()
    upvalue_names = [r.load_string() for _ in range(upvalue_name_size)]

    return {
        "source": source,
        "linedefined": linedefined,
        "lastlinedefined": lastlinedefined,
        "numparams": numparams,
        "is_vararg": is_vararg,
        "maxstacksize": maxstacksize,
        "code": code,
        "constants": constants,
        "upvalues": upvalues,
        "protos": protos,
        "abslineinfo": abslineinfo,
        "locvars": locvars,
        "upvalue_names": upvalue_names,
    }


def decode_instruction(raw: int) -> dict[str, int]:
    return {
        "raw": raw,
        "op": raw & 0x7F,
        "A": (raw >> 7) & 0xFF,
        "k": (raw >> 15) & 1,
        "B": (raw >> 16) & 0xFF,
        "C": (raw >> 24) & 0xFF,
        "Bx": (raw >> 15) & 0x1FFFF,
        "Ax": (raw >> 7) & 0x1FFFFFF,
        "sBx": ((raw >> 15) & 0x1FFFF) - 65535,
    }


def instruction_name(raw: int) -> str:
    op = raw & 0x7F
    return OPNAMES[op] if 0 <= op < len(OPNAMES) else f"OP_{op}"


def read_luac_bytes(data: bytes) -> tuple[dict[str, Any], dict[str, Any]]:
    r = Reader(data)
    header = parse_header(r)
    header["main_upvalue_count"] = r.byte()
    proto = parse_proto(r)
    header["file_size"] = len(data)
    header["parsed_size"] = r.o
    if r.o != len(data):
        header["warning"] = f"parsed={r.o}, file_size={len(data)}"
    return header, proto


def execute_table_chunk(proto: dict[str, Any]) -> dict[str, Any]:
    constants = proto["constants"]
    code = proto["code"]
    regs: list[Any] = [None] * max(512, int(proto.get("maxstacksize") or 0) + 64)
    globals_table: dict[Any, Any] = {}
    returned_values: list[Any] = []
    pc = 0

    def k_or_r(is_const: int, idx: int) -> Any:
        return constants[idx] if is_const else regs[idx]

    while pc < len(code):
        ins = decode_instruction(code[pc])
        name = instruction_name(code[pc])

        if name in {"VARARGPREP", "EXTRAARG"}:
            pc += 1
            continue
        if name == "MOVE":
            regs[ins["A"]] = regs[ins["B"]]
            pc += 1
            continue
        if name == "NEWTABLE":
            regs[ins["A"]] = LuaTable()
            pc += 1
            if pc < len(code) and instruction_name(code[pc]) == "EXTRAARG":
                pc += 1
            continue
        if name == "LOADI":
            regs[ins["A"]] = ins["sBx"]
            pc += 1
            continue
        if name == "LOADF":
            regs[ins["A"]] = float(ins["sBx"])
            pc += 1
            continue
        if name == "LOADK":
            regs[ins["A"]] = constants[ins["Bx"]]
            pc += 1
            continue
        if name == "LOADKX":
            if pc + 1 >= len(code) or instruction_name(code[pc + 1]) != "EXTRAARG":
                raise ValueError("LOADKX without EXTRAARG")
            extra = decode_instruction(code[pc + 1])
            regs[ins["A"]] = constants[extra["Ax"]]
            pc += 2
            continue
        if name == "LOADTRUE":
            regs[ins["A"]] = True
            pc += 1
            continue
        if name == "LOADFALSE":
            regs[ins["A"]] = False
            pc += 1
            continue
        if name == "LOADNIL":
            for idx in range(ins["A"], ins["A"] + ins["B"] + 1):
                regs[idx] = None
            pc += 1
            continue
        if name == "SETFIELD":
            table = regs[ins["A"]]
            if not isinstance(table, LuaTable):
                raise TypeError("SETFIELD target is not table")
            table.set_field(constants[ins["B"]], k_or_r(ins["k"], ins["C"]))
            pc += 1
            continue
        if name == "SETTABLE":
            table = regs[ins["A"]]
            if not isinstance(table, LuaTable):
                raise TypeError("SETTABLE target is not table")
            table.set_field(regs[ins["B"]], k_or_r(ins["k"], ins["C"]))
            pc += 1
            continue
        if name == "SETI":
            table = regs[ins["A"]]
            if not isinstance(table, LuaTable):
                raise TypeError("SETI target is not table")
            table.set_field(ins["B"], k_or_r(ins["k"], ins["C"]))
            pc += 1
            continue
        if name == "SETLIST":
            table = regs[ins["A"]]
            if not isinstance(table, LuaTable):
                raise TypeError("SETLIST target is not table")
            count = ins["B"]
            base = ins["C"]
            if count == 0:
                raise NotImplementedError("SETLIST with variable count is not supported")
            if ins["k"]:
                if pc + 1 >= len(code) or instruction_name(code[pc + 1]) != "EXTRAARG":
                    raise ValueError("SETLIST without EXTRAARG")
                extra = decode_instruction(code[pc + 1])
                base = base + (extra["Ax"] << 8)
                pc += 1
            for i in range(1, count + 1):
                table.set_array(base + i, regs[ins["A"] + i])
            pc += 1
            continue
        if name == "SETTABUP":
            globals_table[constants[ins["B"]]] = k_or_r(ins["k"], ins["C"])
            pc += 1
            continue
        if name.startswith("RETURN"):
            if name == "RETURN":
                count = max(0, ins["B"] - 1)
                returned_values = [regs[ins["A"] + i] for i in range(count)]
            elif name == "RETURN1":
                returned_values = [regs[ins["A"]]]
            break

        raise NotImplementedError(f"unsupported opcode: {name}")

    if globals_table:
        return {str(k): to_python(v) for k, v in globals_table.items()}
    if returned_values:
        return {"return": to_python(returned_values[0])}
    return {}


def get_opcode_names(proto: dict[str, Any]) -> list[str]:
    return sorted(set(instruction_name(raw) for raw in proto["code"]))


def collect_strings(value: Any) -> list[str]:
    out: list[str] = []
    if isinstance(value, str):
        if value and value not in out:
            out.append(value)
    elif isinstance(value, list):
        for item in value:
            for text in collect_strings(item):
                if text not in out:
                    out.append(text)
    elif isinstance(value, dict):
        for item in value.values():
            for text in collect_strings(item):
                if text not in out:
                    out.append(text)
    return out


def collect_proto_strings(proto: dict[str, Any]) -> list[str]:
    out: list[str] = []
    for item in proto.get("constants", []):
        if isinstance(item, str) and item and item not in out:
            out.append(item)
    for child in proto.get("protos", []):
        for text in collect_proto_strings(child):
            if text not in out:
                out.append(text)
    return out


def likely_string_asset(decoded_name: str, strings: list[str]) -> bool:
    if not strings:
        return False
    tokens = set(re.split(r"[_\-.\s]+", decoded_name.upper()))
    if tokens & STRING_ASSET_TOKENS:
        return True
    return len(strings) >= 3 and decoded_name.upper().startswith("LUA_")


def analyze_luac(data: bytes, decoded_name: str) -> dict[str, Any]:
    header, proto = read_luac_bytes(data)
    opcodes = get_opcode_names(proto)
    strings = collect_proto_strings(proto)
    result: dict[str, Any] = {
        "is_luac": True,
        "header": header,
        "opcodes": opcodes,
        "string_count": len(strings),
        "strings": strings,
        "category": "unknown",
        "json_data": None,
        "root_name": "",
        "table_error": "",
    }

    try:
        tables = execute_table_chunk(proto)
        if tables:
            root_name, root_value = next(iter(tables.items()))
            result["category"] = "tables"
            result["root_name"] = root_name
            result["json_data"] = {
                "asset_name": decoded_name,
                "root_name": root_name,
                "data": root_value,
            }
            return result
    except Exception as exc:
        result["table_error"] = f"{type(exc).__name__}: {exc}"

    if likely_string_asset(decoded_name, strings):
        result["category"] = "strings"
        result["root_name"] = "strings"
        result["json_data"] = {
            "asset_name": decoded_name,
            "strings": strings,
        }
    else:
        result["category"] = "script"
    return result


def normalize_textasset_script(script: Any) -> bytes:
    if isinstance(script, bytes):
        return script
    if isinstance(script, bytearray):
        return bytes(script)
    if isinstance(script, str):
        try:
            return script.encode("utf-8", errors="surrogateescape")
        except UnicodeEncodeError:
            return script.encode("utf-8", errors="surrogatepass")
    return bytes(script)


def decrypt_crypto2(script_bytes: bytes) -> bytes:
    buf = bytearray(script_bytes)
    offset = 0
    mask_index = 0
    size = len(buf)

    while offset < size:
        mask = CRYPTO2_MASKS[mask_index]
        remaining = size - offset

        if remaining >= 8:
            value = struct.unpack_from("<Q", buf, offset)[0]
            high = value & 0xFFFFFFFF00000000
            low = value & 0x00000000FFFFFFFF
            low = (
                ((low & 0xFF000000) >> 8)
                | ((low & 0x00FF0000) << 8)
                | ((low & 0x0000FF00) >> 8)
                | ((low & 0x000000FF) << 8)
            )
            value = high | low
            value = ((value & EVEN_MASK) >> 1) | ((value & ODD_MASK) << 1)
            value ^= mask
            struct.pack_into("<Q", buf, offset, value & 0xFFFFFFFFFFFFFFFF)
            offset += 8
        else:
            key_byte = mask & 0xFF
            for i in range(remaining):
                buf[offset + i] ^= key_byte
            offset += remaining

        mask_index = (mask_index + 1) % len(CRYPTO2_MASKS)

    return bytes(buf)


def choose_luac_payload(raw: bytes) -> tuple[bytes | None, str, str]:
    if raw.startswith(b"\x1bLua"):
        return raw, "raw_luac", ""
    first16 = raw[:16].hex()
    decrypted = decrypt_crypto2(raw)
    if decrypted.startswith(b"\x1bLua"):
        return decrypted, "crypto2", ""
    if raw.startswith(bytes.fromhex("f394f36743f58da25c10c853306920bc")):
        return None, "decrypt_failed", f"crypto2 output is not LUAC; encrypted header: {first16}; decrypted header: {decrypted[:16].hex()}"
    return None, "not_luac", f"header: {first16}"


def is_raw_candidate(path: Path, scan_all_raw: bool) -> bool:
    if scan_all_raw:
        return True
    if path.suffix.lower() in RAW_CANDIDATE_EXTS:
        return True
    try:
        return path.read_bytes()[:4] == b"\x1bLua"
    except Exception:
        return False


def output_file_path(
    stage2_dir: Path,
    output_type: str,
    source_folder: str,
    asset_index: int,
    decoded_name: str,
    suffix: str,
    write_mode: str,
) -> tuple[Path | None, Path]:
    type_dir = STAGE2_FORMAT_DIRS[output_type]
    safe_name = sanitize_filename(decoded_name, max_len=140)
    rel = Path(type_dir) / source_folder / f"{asset_index:03d}_{safe_name}{suffix}"
    base_path = stage2_dir / rel
    return output_path_for_write_mode(base_path, write_mode), base_path


def write_json_output(path: Path, data: Any) -> None:
    write_json(path, data)


def write_lua_output(path: Path, decoded_name: str, analysis: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = [
        "-- Best-effort Lua output generated by 02_extract_textasset.py",
        f"-- asset_name: {decoded_name}",
        "",
    ]
    if analysis["category"] in {"tables", "strings"} and analysis.get("json_data") is not None:
        root = analysis.get("root_name") or sanitize_filename(decoded_name)
        data = analysis["json_data"].get("data", analysis["json_data"].get("strings"))
        lines.append(f"{sanitize_filename(root)} = {to_lua_literal(data)}")
    else:
        lines.append("-- Full source decompilation is not available for this script-like bytecode.")
        lines.append("return {")
        lines.append(f"    asset_name = {to_lua_literal(decoded_name)},")
        lines.append(f"    category = {to_lua_literal(analysis.get('category'))},")
        lines.append(f"    opcodes = {to_lua_literal(analysis.get('opcodes', []), 4)},")
        lines.append(f"    strings = {to_lua_literal(analysis.get('strings', []), 4)},")
        lines.append("}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")


def write_luac_output(path: Path, data: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(data)


def xlsx_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    elif not isinstance(value, (str, int, float, bool)) and value is not None:
        value = str(value)
    return clean_excel_value(value)


def write_key_value_rows(ws: Any, data: dict[Any, Any]) -> None:
    ws.append(["key", "value"])
    for key, value in data.items():
        ws.append([xlsx_cell_value(key), xlsx_cell_value(value)])


def write_list_rows(ws: Any, rows: list[Any]) -> None:
    if not rows:
        ws.append(["value"])
        return

    if all(isinstance(item, dict) for item in rows):
        columns: list[str] = []
        for item in rows:
            for key in item.keys():
                key_text = str(key)
                if key_text not in columns:
                    columns.append(key_text)
        ws.append([xlsx_cell_value(col) for col in columns])
        for item in rows:
            ws.append([xlsx_cell_value(item.get(col, "")) for col in columns])
        return

    if all(isinstance(item, (list, tuple)) for item in rows):
        max_len = max((len(item) for item in rows), default=0)
        if max_len == 2:
            ws.append(["key", "text"])
            for item in rows:
                ws.append([xlsx_cell_value(item[0]), xlsx_cell_value(item[1])])
            return

        ws.append(["index", *[f"value_{i}" for i in range(1, max_len + 1)]])
        for row_index, item in enumerate(rows, start=1):
            values = list(item)
            values.extend([""] * (max_len - len(values)))
            ws.append([row_index, *[xlsx_cell_value(value) for value in values]])
        return

    ws.append(["index", "value"])
    for row_index, value in enumerate(rows, start=1):
        ws.append([row_index, xlsx_cell_value(value)])


def write_data_sheet(wb: Any, data: Any) -> None:
    ws = wb.create_sheet("data")

    if isinstance(data, dict) and "strings" in data and isinstance(data["strings"], list):
        ws.append(["index", "text"])
        for row_index, text in enumerate(data["strings"], start=1):
            ws.append([row_index, xlsx_cell_value(text)])
        return

    if isinstance(data, dict) and "data" in data:
        payload = data["data"]
    else:
        payload = data

    if isinstance(payload, list):
        write_list_rows(ws, payload)
    elif isinstance(payload, dict):
        write_key_value_rows(ws, payload)
    else:
        ws.append(["value"])
        ws.append([xlsx_cell_value(payload)])


def write_index_sheet(wb: Any, record: dict[str, Any]) -> None:
    ws = wb.create_sheet("index")
    for key, value in record.items():
        ws.append([xlsx_cell_value(key), xlsx_cell_value(value)])


def write_flat_sheet(wb: Any, data: Any) -> None:
    ws = wb.create_sheet("path_value")
    ws.append(["path", "value_type", "value"])
    for row in iter_flatten_path_value(data):
        ws.append([
            xlsx_cell_value(row["path"]),
            xlsx_cell_value(row["value_type"]),
            xlsx_cell_value(row["value"]),
        ])


def write_xlsx_output(path: Path, record: dict[str, Any], data: Any, include_flat_sheet: bool = False) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")

    wb = Workbook(write_only=True)
    try:
        write_data_sheet(wb, data)
        write_index_sheet(wb, record)
        if include_flat_sheet:
            write_flat_sheet(wb, data)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
    except Exception:
        with suppress(Exception):
            wb.close()
        with suppress(Exception):
            if path.exists():
                path.unlink()
        raise
    finally:
        with suppress(Exception):
            wb.close()


def build_asset_record(
    source_file: Path,
    source_root: Path,
    source_folder: str,
    asset_index: int,
    path_id: Any,
    encrypted_name: str,
    output_type: str,
) -> dict[str, Any]:
    decoded_name, shift, method = best_decode_asset_name(encrypted_name)
    lang = detect_language(decoded_name)
    return {
        "source_file": source_file.name,
        "source_path": str(source_file),
        "source_folder": source_folder,
        "relative_source_path": source_file.relative_to(source_root).as_posix() if source_file.is_relative_to(source_root) else source_file.name,
        "asset_index": asset_index,
        "path_id": path_id,
        "encrypted_asset_name": encrypted_name,
        "decoded_asset_name": decoded_name,
        "decode_shift": shift,
        "decode_method": method,
        **lang,
        "output_type": output_type,
        "category": "unknown",
        "status": "unknown",
        "message": "",
        "file": "",
        "row_count": 0,
    }


def save_output(
    stage2_dir: Path,
    output_type: str,
    record: dict[str, Any],
    analysis: dict[str, Any],
    luac_data: bytes,
    write_mode: str,
    xlsx_flat_sheet: bool,
) -> dict[str, Any]:
    category = str(analysis["category"])
    decoded_name = str(record["decoded_asset_name"])
    source_folder = str(record["source_folder"])
    asset_index = int(record["asset_index"])
    record["category"] = category

    if output_type in {"json", "xlsx"} and category not in {"tables", "strings"}:
        record["status"] = "skipped_unknown_or_script"
        record["message"] = "JSON/XLSX로 변환할 수 있는 정적 구조로 판단하지 못했습니다."
        return record

    def mark_skipped_existing(base_path: Path) -> dict[str, Any]:
        record["status"] = "skipped_existing"
        record["file"] = relative_posix(base_path, stage2_dir)
        record["message"] = "기존 결과 파일이 있어서 저장하지 않았습니다."
        return record

    if output_type == "json":
        path, base_path = output_file_path(stage2_dir, output_type, source_folder, asset_index, decoded_name, ".json", write_mode)
        if path is None:
            return mark_skipped_existing(base_path)
        write_json_output(path, analysis["json_data"])
        record["file"] = relative_posix(path, stage2_dir)
        record["status"] = "ok"
        record["row_count"] = len(analysis.get("strings") or [])
        record["message"] = f"saved json {category}"
        return record

    if output_type == "xlsx":
        path, base_path = output_file_path(stage2_dir, output_type, source_folder, asset_index, decoded_name, ".xlsx", write_mode)
        if path is None:
            return mark_skipped_existing(base_path)
        write_xlsx_output(path, record, analysis["json_data"], include_flat_sheet=xlsx_flat_sheet)
        record["file"] = relative_posix(path, stage2_dir)
        record["status"] = "ok"
        record["row_count"] = len(analysis.get("strings") or [])
        record["xlsx_flat_sheet"] = xlsx_flat_sheet
        record["message"] = f"saved xlsx {category}"
        return record

    if output_type == "lua":
        path, base_path = output_file_path(stage2_dir, output_type, source_folder, asset_index, decoded_name, ".lua", write_mode)
        if path is None:
            return mark_skipped_existing(base_path)
        write_lua_output(path, decoded_name, analysis)
        record["file"] = relative_posix(path, stage2_dir)
        record["status"] = "ok"
        record["row_count"] = len(analysis.get("strings") or [])
        record["message"] = "saved lua best-effort"
        return record

    if output_type == "luac":
        path, base_path = output_file_path(stage2_dir, output_type, source_folder, asset_index, decoded_name, ".luac", write_mode)
        if path is None:
            return mark_skipped_existing(base_path)
        write_luac_output(path, luac_data)
        record["file"] = relative_posix(path, stage2_dir)
        record["status"] = "ok"
        record["row_count"] = len(analysis.get("strings") or [])
        record["message"] = "saved luac"
        return record

    raise ValueError(f"unknown output type: {output_type}")


def process_textasset_payload(
    raw: bytes,
    base_record: dict[str, Any],
    stage2_dir: Path,
    output_type: str,
    write_mode: str,
    xlsx_flat_sheet: bool,
) -> dict[str, Any]:
    luac_data, decode_method, decode_message = choose_luac_payload(raw)
    base_record["payload_decode_method"] = decode_method
    base_record["payload_size"] = len(raw)

    if luac_data is None:
        base_record["status"] = "skipped_not_luac"
        base_record["message"] = decode_message
        return base_record

    try:
        analysis = analyze_luac(luac_data, str(base_record["decoded_asset_name"]))
        base_record["luac_header"] = analysis.get("header", {})
        base_record["opcodes"] = analysis.get("opcodes", [])
        base_record["string_count"] = analysis.get("string_count", 0)
        base_record["table_error"] = analysis.get("table_error", "")
        return save_output(stage2_dir, output_type, base_record, analysis, luac_data, write_mode, xlsx_flat_sheet)
    except Exception as exc:
        base_record["status"] = "error"
        base_record["message"] = f"{type(exc).__name__}: {exc}"
        return base_record


def extract_textassets_from_bundle(
    bundle_path: Path,
    source_root: Path,
    stage2_dir: Path,
    output_type: str,
    write_mode: str,
    xlsx_flat_sheet: bool,
    verbose: bool,
) -> tuple[str, list[dict[str, Any]], str]:
    if UnityPy is None:
        return "error", [], "UnityPy is not installed"

    try:
        env = UnityPy.load(str(bundle_path))
    except Exception as exc:
        return "not_unity_bundle", [], f"{type(exc).__name__}: {exc}"

    source_folder = safe_source_folder_name(bundle_path, source_root)
    records: list[dict[str, Any]] = []
    textasset_index = 0
    for obj_index, obj in enumerate(env.objects):
        try:
            type_name = getattr(getattr(obj, "type", None), "name", "")
            if type_name != "TextAsset":
                continue
            data = obj.read()
            encrypted_name = str(getattr(data, "m_Name", "") or f"textasset_{obj_index}")
            raw = normalize_textasset_script(getattr(data, "m_Script", b""))
            path_id = getattr(obj, "path_id", "")
            record = build_asset_record(
                bundle_path, source_root, source_folder, textasset_index, path_id, encrypted_name, output_type
            )
            records.append(process_textasset_payload(raw, record, stage2_dir, output_type, write_mode, xlsx_flat_sheet))
            textasset_index += 1
        except Exception as exc:
            record = build_asset_record(
                bundle_path, source_root, source_folder, textasset_index, getattr(obj, "path_id", ""), f"textasset_{obj_index}", output_type
            )
            record["status"] = "error"
            record["message"] = f"{type(exc).__name__}: {exc}"
            if verbose:
                record["traceback"] = traceback.format_exc()
            records.append(record)
            textasset_index += 1

    if not records:
        return "no_textasset", [], "UnityFS는 열렸지만 TextAsset을 찾지 못했습니다."
    available = sum(1 for r in records if r.get("status") in {"ok", "skipped_existing"} and r.get("file"))
    saved = sum(1 for r in records if r.get("status") == "ok")
    existing = sum(1 for r in records if r.get("status") == "skipped_existing")
    return ("ok" if available else "no_outputs"), records, f"textassets={len(records)}, saved={saved}, existing={existing}"


def process_raw_file(
    path: Path,
    source_root: Path,
    stage2_dir: Path,
    output_type: str,
    write_mode: str,
    xlsx_flat_sheet: bool,
) -> dict[str, Any]:
    source_folder = safe_source_folder_name(path, source_root)
    record = build_asset_record(path, source_root, source_folder, 0, "", path.stem, output_type)
    return process_textasset_payload(path.read_bytes(), record, stage2_dir, output_type, write_mode, xlsx_flat_sheet)


def iter_scan_files(input_path: Path, output_dir: Path) -> tuple[Path, list[Path]]:
    input_path = input_path.resolve()
    output_dir = output_dir.resolve()
    if input_path.is_file():
        return input_path.parent, [input_path]
    if not input_path.is_dir():
        raise FileNotFoundError(f"입력 파일/폴더를 찾을 수 없습니다: {input_path}")
    files = [p for p in sorted(input_path.rglob("*")) if p.is_file() and not p.resolve().is_relative_to(output_dir)]
    return input_path, files


def should_process(path: Path, input_is_file: bool, all_files: bool) -> bool:
    if input_is_file or all_files:
        return True
    return path.suffix == ""


def build_index(
    stage2_dir: Path,
    format_dir: Path,
    output_type: str,
    input_path: Path,
    args: argparse.Namespace,
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    saved = [r for r in records if r.get("status") == "ok" and r.get("file")]
    available = [r for r in records if r.get("status") in {"ok", "skipped_existing"} and r.get("file")]
    return {
        "tool": "02_extract_textasset.py",
        "script_version": SCRIPT_VERSION,
        "input": str(input_path),
        "stage2_dir": str(stage2_dir),
        "output_dir": str(format_dir),
        "output_type": output_type,
        "all_files": bool(args.all_files),
        "xlsx_flat_sheet": bool(args.xlsx_flat_sheet),
        "write_mode": args.write_mode,
        "saved_count": len(saved),
        "existing_count": sum(1 for r in available if r.get("status") == "skipped_existing"),
        "files": available,
    }


def build_report(
    stage2_dir: Path,
    format_dir: Path,
    input_path: Path,
    output_type: str,
    args: argparse.Namespace,
    sources: list[dict[str, Any]],
) -> dict[str, Any]:
    all_assets = [asset for source in sources for asset in source.get("assets", [])]
    return {
        "tool": "02_extract_textasset.py",
        "script_version": SCRIPT_VERSION,
        "input": str(input_path),
        "stage2_dir": str(stage2_dir),
        "output_dir": str(format_dir),
        "output_type": output_type,
        "all_files": bool(args.all_files),
        "xlsx_flat_sheet": bool(args.xlsx_flat_sheet),
        "write_mode": args.write_mode,
        "cleared_entries": int(getattr(args, "cleared_entries", 0)),
        "quiet": bool(args.quiet),
        "unitypy_available": UnityPy is not None,
        "openpyxl_available": Workbook is not None,
        "total_scan_entries": len(sources),
        "saved_count": sum(1 for a in all_assets if a.get("status") == "ok"),
        "existing_count": sum(1 for a in all_assets if a.get("status") == "skipped_existing"),
        "skipped_count": sum(1 for a in all_assets if str(a.get("status", "")).startswith("skipped")),
        "error_count": sum(1 for a in all_assets if a.get("status") == "error"),
        "sources": sources,
    }


def main() -> int:
    project_root = project_root_from_script(__file__)
    ensure_base_dirs(project_root)

    parser = argparse.ArgumentParser(description="CounterSide TextAsset stage2 extractor")
    parser.add_argument("input", nargs="?", default=str(project_root / "1차 복호화"), help="stage1 folder, Unity bundle, or raw LUAC file")
    parser.add_argument("-o", "--output-dir", default=str(project_root / "2차 복호화(TextAsset)"), help="stage2 output folder")
    parser.add_argument("--output", dest="output_type", choices=["json", "xlsx", "lua", "luac"], default="json", help="output format")
    parser.add_argument("--all-files", action="store_true", help="process files with extensions too")
    parser.add_argument("--xlsx-flat-sheet", action="store_true", help="include flattened path_value sheet in XLSX output")
    parser.add_argument("--write-mode", choices=WRITE_MODES, default="replace", help="replace, skip, or append output files")
    parser.add_argument("--quiet", action="store_true", help="hide per-file progress logs")
    args = parser.parse_args()

    input_path = Path(args.input)
    stage2_dir = Path(args.output_dir)
    output_type = str(args.output_type)
    write_mode = normalize_write_mode(args.write_mode)
    ensure_stage2_dirs(stage2_dir, create_format_dirs=True)
    write_stage2_readme(stage2_dir)
    format_dir = stage2_dir / STAGE2_FORMAT_DIRS[output_type]
    format_dir.mkdir(parents=True, exist_ok=True)
    cleared_entries = 0

    if write_mode == "replace":
        cleared_entries = clear_directory_contents(
            format_dir,
            allowed_parent=stage2_dir,
            protected_paths=[
                stage2_dir,
                stage2_dir / "언어별 구분",
                input_path,
            ],
        )
    args.write_mode = write_mode
    args.cleared_entries = cleared_entries

    source_root, files = iter_scan_files(input_path, stage2_dir)
    input_is_file = input_path.is_file()
    sources: list[dict[str, Any]] = []
    saved_records: list[dict[str, Any]] = []
    log_progress = not args.quiet

    if log_progress:
        print("2차 복호화(TextAsset) 작업을 시작합니다.")
        print(f"입력: {input_path}")
        print(f"출력: {stage2_dir}")
        print(f"출력 형식: {output_type}")
        print(f"저장 방식: {write_mode}")
        if output_type == "xlsx":
            print(f"XLSX 펼침 시트: {'사용' if args.xlsx_flat_sheet else '사용 안 함'}")
        if write_mode == "replace":
            print(f"기존 출력 정리: {format_dir} / {cleared_entries}개 항목 삭제")
        print(f"스캔한 파일 수: {len(files)}개")
        print("-" * 60)

    for file_index, path in enumerate(files, start=1):
        rel = path.relative_to(source_root).as_posix() if path.is_relative_to(source_root) else path.name
        source_row: dict[str, Any] = {
            "relative_path": rel,
            "source_path": str(path),
            "status": "",
            "message": "",
            "assets": [],
        }

        if not should_process(path, input_is_file, args.all_files):
            source_row["status"] = "skipped_extension"
            source_row["message"] = "확장자 있는 파일이라 기본 2차 처리 대상에서 제외했습니다."
            sources.append(source_row)
            if log_progress and (file_index == 1 or file_index % 25 == 0 or file_index == len(files)):
                print(f"[{file_index}/{len(files)}] skip extension: {rel}")
            continue

        if log_progress:
            print(f"[{file_index}/{len(files)}] 처리 중: {rel}", flush=True)

        status, records, message = extract_textassets_from_bundle(
            path, source_root, stage2_dir, output_type, write_mode, args.xlsx_flat_sheet, log_progress
        )
        if status == "not_unity_bundle" and is_raw_candidate(path, args.all_files):
            try:
                records = [process_raw_file(path, source_root, stage2_dir, output_type, write_mode, args.xlsx_flat_sheet)]
                status = "ok" if records[0].get("status") in {"ok", "skipped_existing"} else "no_outputs"
                message = str(records[0].get("message") or records[0].get("status"))
            except Exception as exc:
                status = "error"
                records = []
                message = f"{type(exc).__name__}: {exc}"

        source_row["status"] = status
        source_row["message"] = message
        source_row["assets"] = records
        sources.append(source_row)
        saved_records.extend(r for r in records if r.get("status") in {"ok", "skipped_existing"} and r.get("file"))

        if log_progress:
            saved_here = sum(1 for r in records if r.get("status") == "ok" and r.get("file"))
            skipped_here = sum(1 for r in records if str(r.get("status", "")).startswith("skipped"))
            error_here = sum(1 for r in records if r.get("status") == "error")
            print(
                f"  -> {status}: 저장 {saved_here}개, 스킵 {skipped_here}개, 오류 {error_here}개 / {message}",
                flush=True,
            )

    index = build_index(stage2_dir, format_dir, output_type, input_path, args, saved_records)
    report = build_report(stage2_dir, format_dir, input_path, output_type, args, sources)
    index_path = format_dir / "index.json"
    report_path = format_dir / "stage2_report.json"
    write_json(index_path, index)
    write_json(report_path, report)

    if log_progress:
        print("-" * 60)
    print(f"완료: {report_path}")
    print(f"저장한 파일 수: {index['saved_count']}개")
    if index.get("existing_count"):
        print(f"기존 유지 파일 수: {index['existing_count']}개")
    print(f"index: {index_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
